# process_excel.py
import sys
from openpyxl import load_workbook
import pandas as pd
import math
import json

def process_excel(file_path):
    try:
        # Load the Excel file
        workbook = load_workbook(filename=file_path)

        # Choose the specific sheet you want to work with
        sheet_name = 'Center - 2'  # Replace with the actual sheet name
        sheet = workbook[sheet_name]

        # Get the maximum row number in the sheet
        max_row = sheet.max_row

        # Define the column letter you want to read
        column_name = 'D'

        # Load the Excel file into a DataFrame
        df = pd.read_excel(file_path)

        # Replace 'D' with the correct column name and get unique values
        svc_type_col = 'Unnamed: 3'
        if svc_type_col in df.columns:
            column_values = df[svc_type_col].unique()

            # Convert the unique values to a Python list if needed
            unique_values_list = column_values.tolist()

        else:
            return "Error: Column not found in the DataFrame."

        cleaned_list = [x for x in unique_values_list if not (isinstance(x, float) and math.isnan(x)) and x != 'Description']

        # Create a dictionary with values initialized to 0
        svcChargeDict = {item: 0 for item in cleaned_list}

        # Add the 'total' key with a value of 0
        svcChargeDict['Total'] = 0

        # Iterate through rows in the specified column, starting from row 2
        for row_number in range(2, 500):
            cell_value = sheet[f'{column_name}{row_number}'].value
            if isinstance(cell_value, str):
                if cell_value in svcChargeDict:
                    svc_value = sheet[f'{"C"}{row_number}'].value
                    if isinstance(svc_value, int):
                        svcChargeDict[cell_value] += svc_value

        combined_charges_dict = {}

        # Iterate through the original dictionary
        for key, value in svcChargeDict.items():
            base_key = key.split('*')[0].strip()  # Get the base key without anything after '*'
            if base_key not in combined_charges_dict:
                combined_charges_dict[base_key] = value
            else:
                combined_charges_dict[base_key] += value

        total_sum = sum(value for key, value in combined_charges_dict.items() if key != 'Total')

        # Update the 'Total' key with the sum
        combined_charges_dict['Total'] = total_sum

        # Close the workbook
        workbook.close()

        # Convert the dictionary to a JSON string and print each JSON string on a new line
        result_json = json.dumps(combined_charges_dict)
        print(result_json)

    except Exception as e:
        print(f"Error: {str(e)}")

if __name__ == "__main__":
    if len(sys.argv) != 2:
        print(f"Usage: {sys.argv[0]} <file_path>")
        sys.exit(1)

    file_path = sys.argv[1]
    process_excel(file_path)
